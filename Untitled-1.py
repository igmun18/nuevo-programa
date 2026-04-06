import customtkinter as ctk
from tkinterdnd2 import DND_FILES, TkinterDnD
from tkinter import filedialog, messagebox
from PIL import Image
import os
import openpyxl
import re
from openpyxl import load_workbook
from copy import copy
from dateutil.relativedelta import relativedelta
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def obtener_rango_combinado(ws, fila, columna):
    """
    Devuelve el rango combinado al que pertenece una celda (fila, columna).
    Si no pertenece a ninguno, devuelve None.
    """

    # 🛡️ validación inicial
    if fila is None or columna is None:
        return None

    for rango in ws.merged_cells.ranges:
        try:
            min_row = rango.min_row
            max_row = rango.max_row
            min_col = rango.min_col
            max_col = rango.max_col

            # 🛡️ evitar rangos corruptos o incompletos
            if None in (min_row, max_row, min_col, max_col):
                continue

            # 🔍 chequeo de pertenencia
            if (min_row <= fila <= max_row) and (min_col <= columna <= max_col):
                return rango

        except Exception:
            # ignorar rangos problemáticos sin romper el flujo
            continue

    return None

class App(ctk.CTk, TkinterDnD.Tk):
    def __init__(self):

        ctk.CTk.__init__(self)
        TkinterDnD.Tk.__init__(self)

        ctk.set_appearance_mode("dark")
        self.title("Mi primer programa c:")
        self.geometry("700x700")
        self.configure(fg_color="#242424")

        # ---------------- ICONO ----------------
        try:
            self.icono_excel = ctk.CTkImage(
                light_image=Image.open("excel_icon.png"),
                dark_image=Image.open("excel_icon.png"),
                size=(40, 40)
            )
        except:
            self.icono_excel = None

        self.rutas_archivos = []

        # ---------------- CONTENEDOR PRINCIPAL ----------------
        self.container = ctk.CTkFrame(self, fg_color="transparent")
        self.container.pack(fill="both", expand=True, padx=20, pady=20)

        # ---------------- SECCIÓN BUSCAR ----------------
        frame_path = ctk.CTkFrame(self.container)
        frame_path.pack(fill="x", pady=10)

        btn_buscar = ctk.CTkButton(frame_path, text="Buscar", width=100, command=self.buscar_archivo)
        btn_buscar.pack(side="left", padx=10)

        self.entry_path = ctk.CTkEntry(frame_path, placeholder_text="Ruta del archivo...")
        self.entry_path.pack(side="left", padx=10, expand=True, fill="x")

        # ---------------- SECCIÓN PORCENTAJE ----------------
        frame_porcentaje = ctk.CTkFrame(self.container)
        frame_porcentaje.pack(fill="x", pady=10)

        label = ctk.CTkLabel(frame_porcentaje, text="Porcentaje de aumento (%):")
        label.pack(side="left", padx=10)

        self.entry_porcentaje = ctk.CTkEntry(frame_porcentaje, width=100)
        self.entry_porcentaje.insert(0, "15")
        self.entry_porcentaje.pack(side="left", padx=10)

        # ---------------- DRAG & DROP ----------------
        self.frame_drop = ctk.CTkFrame(
            self.container,
            border_width=2,
            border_color="#3B3B3B"
        )
        self.frame_drop.pack(fill="both", expand=True, pady=10)

        self.label_drop = ctk.CTkLabel(
            self.frame_drop,
            text="Arrastra archivos Excel aquí o usa 'Buscar'"
        )
        self.label_drop.pack(pady=10)

        # SCROLLABLE FRAME
        self.lista_iconos_frame = ctk.CTkScrollableFrame(
            self.frame_drop,
            label_text="Archivos seleccionados"
        )
        self.lista_iconos_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # ---------------- BOTÓN ----------------
        self.btn_ejecutar = ctk.CTkButton(
            self.container,
            text="APLICAR INCREMENTO",
            fg_color="#0B1575",
            hover_color="#2808B4",
            height=45,
            command=self.ejecutar_macro
        )
        self.btn_ejecutar.pack(fill="x", pady=15)

        # DRAG & DROP
        self.frame_drop.drop_target_register(DND_FILES)
        self.frame_drop.dnd_bind('<<Drop>>', self.al_soltar_archivo)

    # ---------------- FUNCIONES ----------------

    def buscar_archivo(self):
        archivos = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx *.xls")])
        if archivos:
            self.agregar_a_lista(archivos)

    def al_soltar_archivo(self, event):
        rutas = self.tk.splitlist(event.data)
        self.agregar_a_lista(rutas)

    def agregar_a_lista(self, rutas):
        for ruta in rutas:
            ruta = ruta.strip('{}')
            if ruta not in self.rutas_archivos and ruta.lower().endswith(('.xlsx', '.xls')):
                self.rutas_archivos.append(ruta)

        self.renderizar_iconos()

    # ---------------- GRILLA DE ICONOS ----------------

    def renderizar_iconos(self):
        # limpiar todo
        for widget in self.lista_iconos_frame.winfo_children():
            widget.destroy()

        columnas = 4

        for index, ruta in enumerate(self.rutas_archivos):
            nombre = os.path.basename(ruta)

            fila = index // columnas
            col = index % columnas

            item = ctk.CTkFrame(self.lista_iconos_frame, width=120, height=120)
            item.grid(row=fila, column=col, padx=10, pady=10)
            item.grid_propagate(False)

            # icono
            label_img = ctk.CTkLabel(
                item,
                image=self.icono_excel,
                text="" if self.icono_excel else "📄"
            )
            label_img.pack(pady=(10, 5))

            # nombre corto
            nombre_corto = nombre[:12] + "..." if len(nombre) > 12 else nombre
            label_nombre = ctk.CTkLabel(item, text=nombre_corto)
            label_nombre.pack()

            # botón eliminar
            btn_eliminar = ctk.CTkButton(
                item,
                text="✕",
                width=20,
                height=20,
                fg_color="#A12121",
                hover_color="#E63946",
                command=lambda r=ruta: self.eliminar_archivo(r)
            )
            btn_eliminar.place(relx=1, rely=0, anchor="ne")

    def eliminar_archivo(self, ruta):
        if ruta in self.rutas_archivos:
            self.rutas_archivos.remove(ruta)

        self.renderizar_iconos()

    # ---------------- MACRO ----------------
    

    def ejecutar_macro(self):

        try:
            factor = 1 + (float(self.entry_porcentaje.get()) / 100)
        except ValueError:
            messagebox.showerror("Error", "Ingresa un porcentaje válido.")
            return

        if not self.rutas_archivos:
            messagebox.showwarning("Atención", "No hay archivos seleccionados.")
            return

        exitos = 0

        for ruta in self.rutas_archivos:
            try:
                wb = openpyxl.load_workbook(ruta, data_only=True)

                for ws in wb.worksheets:

                    fila_header = None

                    # 🔍 BUSCAR HEADER (fila que contiene fecha)
                    for fila in range(1, 15):
                        for col in range(ws.max_column, 0, -1):

                            valor = ws.cell(row=fila, column=col).value

                            if isinstance(valor, datetime):
                                fila_header = fila
                                break

                            elif isinstance(valor, str):
                                try:
                                    datetime.strptime(valor, "%d/%m/%Y")
                                    fila_header = fila
                                    break
                                except:
                                    pass

                        if fila_header:
                            break

                    if not fila_header:
                        print(f"⚠️ Hoja ignorada: {ws.title}")
                        continue

                    # 🔥 BUSCAR ÚLTIMA COLUMNA CON FECHA REAL
                    ultima_columna = None

                    for col in range(ws.max_column, 0, -1):
                        valor = ws.cell(row=fila_header, column=col).value

                        if isinstance(valor, datetime):
                            ultima_columna = col
                            break

                        elif isinstance(valor, str):
                            try:
                                datetime.strptime(valor, "%d/%m/%Y")
                                ultima_columna = col
                                break
                            except:
                                pass

                    if ultima_columna is None:
                        print(f"⚠️ No se encontró columna válida en {ws.title}")
                        continue

                    # 🔥 DETECTAR BLOQUE REAL (merge)
                    rango_header = obtener_rango_combinado(ws, fila_header, ultima_columna)

                    if rango_header and rango_header.min_row == rango_header.max_row:
                        col_inicio = rango_header.min_col
                        col_fin = rango_header.max_col
                    else:
                        col_inicio = ultima_columna
                        col_fin = ultima_columna

                    cantidad_columnas = col_fin - col_inicio + 1
                    nueva_col_inicio = col_fin + 1

                    # 📅 FECHA BASE (SIEMPRE desde la primera del merge)
                    header_base = ws.cell(row=fila_header, column=col_inicio)
                    fecha = header_base.value

                    if isinstance(fecha, str):
                        try:
                            fecha = datetime.strptime(fecha, "%d/%m/%Y")
                        except:
                            fecha = None

                    elif isinstance(fecha, (int, float)):
                        from openpyxl.utils.datetime import from_excel
                        fecha = from_excel(fecha)

                    if not fecha:
                        continue

                    nueva_fecha = fecha + relativedelta(months=1)

                    # 🧱 COPIAR HEADER COMPLETO
                    for i in range(cantidad_columnas):
                        col_origen = col_inicio + i
                        col_destino = nueva_col_inicio + i

                        header_origen = ws.cell(row=fila_header, column=col_origen)
                        header_destino = ws.cell(row=fila_header, column=col_destino)

                        header_destino.value = header_origen.value

                        header_destino.font = copy(header_origen.font)
                        header_destino.border = copy(header_origen.border)
                        header_destino.fill = copy(header_origen.fill)
                        header_destino.alignment = copy(header_origen.alignment)

                    # 📅 nueva fecha solo en la primera
                    ws.cell(row=fila_header, column=nueva_col_inicio).value = nueva_fecha
                    ws.cell(row=fila_header, column=nueva_col_inicio).number_format = "DD/MM/YYYY"

                    # 🔗 recrear merge
                    if cantidad_columnas > 1:
                        ws.merge_cells(
                            start_row=fila_header,
                            end_row=fila_header,
                            start_column=nueva_col_inicio,
                            end_column=nueva_col_inicio + cantidad_columnas - 1
                        )

                    # 🔢 DATOS
                    for fila in range(fila_header + 1, ws.max_row + 1):
                        for i in range(cantidad_columnas):

                            col_origen = col_inicio + i
                            col_destino = nueva_col_inicio + i

                            celda_origen = ws.cell(row=fila, column=col_origen)
                            celda_destino = ws.cell(row=fila, column=col_destino)

                            valor = celda_origen.value

                            if isinstance(valor, (int, float)):
                                celda_destino.value = round(valor * factor,2)

                            elif isinstance(valor, str):
                                match = re.search(r'(NN\s*[xX]\s*)(\d+)', valor)

                                if match:
                                    prefijo = match.group(1)
                                    numero = match.group(2)
                                    celda_destino.value = f"{prefijo}{round(float(numero) * factor,2)}"
                                else:
                                    celda_destino.value = valor
                            else:
                                celda_destino.value = valor

                            if celda_origen.has_style:
                                celda_destino.font = copy(celda_origen.font)
                                celda_destino.border = copy(celda_origen.border)
                                celda_destino.fill = copy(celda_origen.fill)
                                celda_destino.number_format = copy(celda_origen.number_format)
                                celda_destino.alignment = copy(celda_origen.alignment)

                    # 📏 ancho columnas
                    for i in range(cantidad_columnas):
                        col_origen = get_column_letter(col_inicio + i)
                        col_destino = get_column_letter(nueva_col_inicio + i)

                        ws.column_dimensions[col_destino].width = ws.column_dimensions[col_origen].width

                wb.save(ruta)
                exitos += 1

            except Exception as e:
                print(f"Error en {ruta}: {e}")

        messagebox.showinfo("Hecho", f"Se procesaron {exitos} archivos correctamente.")


if __name__ == "__main__": 
    app = App() 
    app.mainloop()